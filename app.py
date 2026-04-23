from __future__ import annotations

import argparse
import mimetypes
import sys
import time
from pathlib import Path
from typing import Generator, Union
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import urlopen

from openpyxl import load_workbook


DEFAULT_EXCEL_PATH = "SP.xlsx"
DEFAULT_OUTPUT_DIR = "downloaded_images"
CODE_HEADER = "Mã hàng"
IMAGE_HEADER = "Hình ảnh (url1,url2...)"
REQUEST_TIMEOUT_SECONDS = 30


def application_dir() -> Path:
    if getattr(sys, "frozen", False):
        executable_path = Path(sys.executable).resolve()
        for parent in executable_path.parents:
            if parent.suffix.lower() == ".app":
                return parent.parent
        return executable_path.parent
    return Path.cwd()


def resolve_user_path(path_value: str) -> Path:
    path = Path(path_value)
    if path.is_absolute():
        return path
    return application_dir() / path


def can_prompt_user() -> bool:
    stdin = sys.stdin
    return bool(stdin and hasattr(stdin, "isatty") and stdin.isatty())


def show_message(title: str, message: str, *, error: bool = False) -> None:
    try:
        import tkinter
        from tkinter import messagebox

        root = tkinter.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        if error:
            messagebox.showerror(title, message, parent=root)
        else:
            messagebox.showinfo(title, message, parent=root)
        root.destroy()
    except Exception:
        stream = sys.stderr if error else sys.stdout
        print(f"{title}: {message}", file=stream)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download product images from an Excel file."
    )
    parser.add_argument(
        "excel_path",
        nargs="?",
        default=DEFAULT_EXCEL_PATH,
        help=f"Path to the Excel file. Defaults to {DEFAULT_EXCEL_PATH}.",
    )
    parser.add_argument(
        "output_dir",
        nargs="?",
        default=DEFAULT_OUTPUT_DIR,
        help=f"Directory to save images. Defaults to {DEFAULT_OUTPUT_DIR}.",
    )
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="Force CLI mode instead of GUI.",
    )
    return parser.parse_args()


def wait_for_exit(message: str = "Press ENTER to exit.") -> None:
    if not can_prompt_user():
        return
    try:
        input(message)
    except (EOFError, RuntimeError):
        pass


def fatal(message: str) -> int:
    print(f"Error: {message}", file=sys.stderr)
    print()
    if not can_prompt_user():
        show_message("Download failed", message, error=True)
        return 1
    wait_for_exit()
    return 1


def sanitize_filename(value: str) -> str:
    cleaned = "".join("_" if char in '<>:"/\\|?*' else char for char in value.strip())
    cleaned = cleaned.rstrip(". ")
    return cleaned or "unknown"


def split_image_urls(value: object) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split(",") if part and part.strip()]


def load_rows(excel_path: Path) -> tuple[Generator, int, int]:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Unable to read Excel file '{excel_path}': {exc}") from exc

    if not workbook.sheetnames:
        raise ValueError(f"Workbook '{excel_path}' does not contain any worksheets.")

    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = list(next(rows))
    except StopIteration as exc:
        raise ValueError(f"Workbook '{excel_path}' is empty.") from exc

    missing_headers = [
        header for header in (CODE_HEADER, IMAGE_HEADER) if header not in header_row
    ]
    if missing_headers:
        joined = ", ".join(missing_headers)
        raise ValueError(f"Missing required column(s): {joined}")

    return rows, header_row.index(CODE_HEADER), header_row.index(IMAGE_HEADER)


def guess_extension(url: str, content_type: str | None) -> str:
    path = urlparse(url).path
    suffix = Path(path).suffix
    if suffix:
        return suffix.lower()

    if content_type:
        guessed = mimetypes.guess_extension(content_type.split(";", 1)[0].strip())
        if guessed:
            return guessed.lower()

    return ".jpg"


def choose_target_path(output_dir: Path, base_name: str, extension: str) -> Path:
    candidate = output_dir / f"{base_name}{extension}"
    counter = 1
    while candidate.exists():
        candidate = output_dir / f"{base_name}_{counter}{extension}"
        counter += 1
    return candidate


def build_base_name(product_code: str, image_index: int, total_images: int) -> str:
    safe_code = sanitize_filename(product_code)
    if total_images > 1:
        return f"{safe_code}_{image_index}"
    return safe_code


def download_file(url: str, destination: Path) -> None:
    with urlopen(url, timeout=REQUEST_TIMEOUT_SECONDS) as response:
        extension = guess_extension(url, response.headers.get_content_type())
        target_path = choose_target_path(destination.parent, destination.stem, extension)
        with target_path.open("wb") as output_file:
            output_file.write(response.read())


def prompt_with_default(label: str, default_value: str) -> str:
    if not can_prompt_user():
        return default_value
    prompt = f"{label} [{default_value}]: " if default_value else f"{label}: "
    try:
        entered_value = input(prompt).strip()
    except (EOFError, RuntimeError):
        return default_value
    return entered_value or default_value


def confirm_paths(excel_path: Path, output_dir: Path) -> tuple[Path, Path]:
    if not can_prompt_user():
        output_dir.mkdir(parents=True, exist_ok=True)
        return excel_path, output_dir

    print("Product image downloader")
    print()

    selected_excel_path = excel_path
    while True:
        selected_excel_path = Path(
            prompt_with_default("Input Excel file", str(selected_excel_path))
        )

        if selected_excel_path.exists():
            break

        print(f"Warning: input file not found: {selected_excel_path}")
        print("Please press ENTER to keep the same name or type a different file path.")
        print()

    selected_output_dir = Path(
        prompt_with_default("Output directory", str(output_dir))
    )

    if selected_output_dir.exists():
        print(f"Output directory ready: {selected_output_dir.resolve()}")
    else:
        selected_output_dir.mkdir(parents=True, exist_ok=True)
        print(f"Created output directory: {selected_output_dir.resolve()}")

    print()
    wait_for_exit("Press ENTER to download.")
    return selected_excel_path, selected_output_dir


DownloadResult = tuple[str, str, Union[str, None]]


def run_downloads(
    excel_path: Path,
    output_dir: Path,
    log_callback: callable | None = None,
    progress_callback: callable | None = None,
) -> Generator[DownloadResult, None, tuple[int, int, int]]:
    rows, code_index, image_index = load_rows(excel_path)

    output_dir.mkdir(parents=True, exist_ok=True)

    processed_rows = 0
    success_count = 0
    failure_count = 0
    total_images = 0

    for row in rows:
        processed_rows += 1

    rows, code_index, image_index = load_rows(excel_path)

    for row_number, row in enumerate(rows, start=2):
        product_code_raw = row[code_index] if code_index < len(row) else None
        image_urls_raw = row[image_index] if image_index < len(row) else None

        product_code = "" if product_code_raw is None else str(product_code_raw).strip()
        image_urls = split_image_urls(image_urls_raw)

        if not product_code:
            msg = f"Warning: row {row_number} skipped because '{CODE_HEADER}' is empty."
            if log_callback:
                log_callback(msg, "warning")
            yield "info", product_code, msg
            continue

        if not image_urls:
            msg = (
                f"Warning: row {row_number} ({product_code}) skipped because "
                f"'{IMAGE_HEADER}' is empty."
            )
            if log_callback:
                log_callback(msg, "warning")
            yield "info", product_code, msg
            continue

        for image_number, url in enumerate(image_urls, start=1):
            base_name = build_base_name(product_code, image_number, len(image_urls))
            destination = output_dir / base_name
            try:
                download_file(url, destination)
                success_count += 1
                msg = f"Downloaded: {product_code} <- {url}"
                if log_callback:
                    log_callback(msg, "success")
                yield "success", product_code, msg
            except (HTTPError, URLError, TimeoutError, ValueError, OSError) as exc:
                failure_count += 1
                msg = f"Failed: {product_code} <- {url} ({exc})"
                if log_callback:
                    log_callback(msg, "error")
                yield "error", product_code, msg

            if progress_callback:
                progress_callback(success_count + failure_count, total_images)

    yield "info", "", f"Done: {success_count} success, {failure_count} failed"


def count_image_urls(excel_path: Path) -> int:
    rows, code_index, image_index = load_rows(excel_path)
    total = 0
    for row in rows:
        product_code_raw = row[code_index] if code_index < len(row) else None
        image_urls_raw = row[image_index] if image_index < len(row) else None
        product_code = "" if product_code_raw is None else str(product_code_raw).strip()
        image_urls = split_image_urls(image_urls_raw)
        if product_code and image_urls:
            total += len(image_urls)
    return total


def main() -> int:
    args = parse_args()

    if not args.no_gui and not can_prompt_user():
        from PyQt6.QtWidgets import QApplication

        app = QApplication(sys.argv)
        return gui_main(app)

    excel_path = resolve_user_path(args.excel_path)
    output_dir = resolve_user_path(args.output_dir)

    excel_path, output_dir = confirm_paths(excel_path, output_dir)

    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        rows, code_index, image_index = load_rows(excel_path)
    except (FileNotFoundError, RuntimeError, ValueError) as exc:
        return fatal(str(exc))

    processed_rows = 0
    success_count = 0
    failure_count = 0

    for row_number, row in enumerate(rows, start=2):
        processed_rows += 1

        product_code_raw = row[code_index] if code_index < len(row) else None
        image_urls_raw = row[image_index] if image_index < len(row) else None

        product_code = "" if product_code_raw is None else str(product_code_raw).strip()
        image_urls = split_image_urls(image_urls_raw)

        if not product_code:
            print(f"Warning: row {row_number} skipped because '{CODE_HEADER}' is empty.")
            continue

        if not image_urls:
            print(
                f"Warning: row {row_number} ({product_code}) skipped because "
                f"'{IMAGE_HEADER}' is empty."
            )
            continue

        for image_number, url in enumerate(image_urls, start=1):
            base_name = build_base_name(product_code, image_number, len(image_urls))
            destination = output_dir / base_name
            try:
                download_file(url, destination)
                success_count += 1
                print(f"Downloaded: {product_code} <- {url}")
            except (HTTPError, URLError, TimeoutError, ValueError, OSError) as exc:
                failure_count += 1
                print(f"Failed: {product_code} <- {url} ({exc})")

    summary = "\n".join(
        [
            "Download summary",
            f"Rows processed: {processed_rows}",
            f"Images downloaded: {success_count}",
            f"Images failed: {failure_count}",
            f"Output directory: {output_dir.resolve()}",
        ]
    )

    print()
    print(summary)
    print()

    if not can_prompt_user():
        show_message("Download complete", summary)
        return 0

    wait_for_exit()

    return 0


def gui_main(app: QApplication) -> int:
    from PyQt6.QtCore import QThread, pyqtSignal, Qt
    from PyQt6.QtGui import QColor, QDragEnterEvent, QDropEvent, QFont
    from PyQt6.QtWidgets import (
        QAbstractItemView,
        QApplication,
        QDialog,
        QFrame,
        QHBoxLayout,
        QLabel,
        QLineEdit,
        QPushButton,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )

    class DownloadThread(QThread):
        log_signal = pyqtSignal(str, str)
        progress_signal = pyqtSignal(int, int)
        finished_signal = pyqtSignal(int, int, int)

        def __init__(self, excel_path: Path, output_dir: Path):
            super().__init__()
            self.excel_path = excel_path
            self.output_dir = output_dir
            self._stop_requested = False

        def run(self):
            success_count = 0
            failure_count = 0
            total_images = count_image_urls(self.excel_path)

            self.output_dir.mkdir(parents=True, exist_ok=True)

            rows, code_index, image_index = load_rows(self.excel_path)

            for row_number, row in enumerate(rows, start=2):
                if self._stop_requested:
                    break

                product_code_raw = row[code_index] if code_index < len(row) else None
                image_urls_raw = row[image_index] if image_index < len(row) else None

                product_code = "" if product_code_raw is None else str(product_code_raw).strip()
                image_urls = split_image_urls(image_urls_raw)

                if not product_code:
                    msg = f"Warning: row {row_number} skipped because '{CODE_HEADER}' is empty."
                    self.log_signal.emit(msg, "warning")
                    continue

                if not image_urls:
                    msg = (
                        f"Warning: row {row_number} ({product_code}) skipped because "
                        f"'{IMAGE_HEADER}' is empty."
                    )
                    self.log_signal.emit(msg, "warning")
                    continue

                for image_number, url in enumerate(image_urls, start=1):
                    if self._stop_requested:
                        break

                    base_name = build_base_name(product_code, image_number, len(image_urls))
                    destination = self.output_dir / base_name
                    try:
                        download_file(url, destination)
                        success_count += 1
                        msg = f"Downloaded: {product_code} <- {url}"
                        self.log_signal.emit(msg, "success")
                    except (HTTPError, URLError, TimeoutError, ValueError, OSError) as exc:
                        failure_count += 1
                        msg = f"Failed: {product_code} <- {url} ({exc})"
                        self.log_signal.emit(msg, "error")

                    self.progress_signal.emit(success_count + failure_count, total_images)

            self.finished_signal.emit(success_count, failure_count, total_images)

        def stop(self):
            self._stop_requested = True

    class DropZone(QFrame):
        def __init__(self, parent: QWidget = None):
            super().__init__(parent)
            self.setAcceptDrops(True)
            self.setMinimumHeight(120)
            self.setFrameStyle(QFrame.Shape.Box | QFrame.Shadow.Raised)
            self.setStyleSheet("""
                QFrame {
                    border: 2px dashed #888;
                    border-radius: 8px;
                    background-color: #f5f5f5;
                }
                QFrame:hover {
                    border-color: #444;
                    background-color: #e8e8e8;
                }
            """)

            layout = QVBoxLayout(self)
            self.label = QLabel("DRAG & DROP EXCEL FILE HERE\n(.xlsx only)", self)
            self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.label.setStyleSheet("font-size: 14px; color: #666;")
            layout.addWidget(self.label)

            self._highlight = False

        def dragEnterEvent(self, event: QDragEnterEvent):
            if event.mimeData().hasUrls():
                urls = event.mimeData().urls()
                if urls and urls[0].toLocalFile().lower().endswith(".xlsx"):
                    self._highlight = True
                    self.setStyleSheet("""
                        QFrame {
                            border: 2px solid #2196F3;
                            border-radius: 8px;
                            background-color: #e3f2fd;
                        }
                    """)
                    event.acceptProposedAction()
                    return
            event.ignore()

        def dragLeaveEvent(self, event):
            self._highlight = False
            self.setStyleSheet("""
                QFrame {
                    border: 2px dashed #888;
                    border-radius: 8px;
                    background-color: #f5f5f5;
                }
                QFrame:hover {
                    border-color: #444;
                    background-color: #e8e8e8;
                }
            """)

        def dropEvent(self, event: QDropEvent):
            self._highlight = False
            self.setStyleSheet("""
                QFrame {
                    border: 2px dashed #888;
                    border-radius: 8px;
                    background-color: #f5f5f5;
                }
                QFrame:hover {
                    border-color: #444;
                    background-color: #e8e8e8;
                }
            """)
            urls = event.mimeData().urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if file_path.lower().endswith(".xlsx"):
                    event.acceptProposedAction()
                    if self.parent():
                        self.parent().on_file_dropped(file_path)

    class MainWindow(QDialog):
        def __init__(self):
            super().__init__()
            self.download_thread = None
            self.start_time = None
            self.total_images = 0
            self.init_ui()

        def init_ui(self):
            self.setWindowTitle("kv-image-downloader")
            self.setMinimumWidth(600)
            self.setMinimumHeight(500)

            main_layout = QVBoxLayout(self)
            main_layout.setSpacing(12)

            self.drop_zone = DropZone(self)
            main_layout.addWidget(self.drop_zone)

            self.file_label = QLabel("Selected file: (none)", self)
            main_layout.addWidget(self.file_label)

            self.output_label = QLabel("Output folder: downloaded_images/", self)
            main_layout.addWidget(self.output_label)

            button_layout = QHBoxLayout()
            self.browse_button = QPushButton("Browse...", self)
            self.browse_button.clicked.connect(self.on_browse)
            button_layout.addWidget(self.browse_button)

            self.clear_button = QPushButton("Clear", self)
            self.clear_button.clicked.connect(self.on_clear)
            self.clear_button.setEnabled(False)
            button_layout.addWidget(self.clear_button)

            button_layout.addStretch()
            main_layout.addLayout(button_layout)

            self.start_button = QPushButton("▶  START DOWNLOAD", self)
            self.start_button.clicked.connect(self.on_start)
            self.start_button.setEnabled(False)
            self.start_button.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                    padding: 12px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #888888;
                }
            """)
            main_layout.addWidget(self.start_button)

            log_label = QLabel("LOG VIEWER", self)
            main_layout.addWidget(log_label)

            self.log_viewer = QTextEdit(self)
            self.log_viewer.setReadOnly(True)
            self.log_viewer.setFont(QFont("monospace", 10))
            self.log_viewer.setStyleSheet("background-color: #1e1e1e; color: #d4d4d4;")
            self.log_viewer.setMinimumHeight(200)
            main_layout.addWidget(self.log_viewer)

            self.status_label = QLabel("Status: Ready", self)
            main_layout.addWidget(self.status_label)

        def on_file_dropped(self, file_path: str):
            from pathlib import Path

            path = Path(file_path)
            if path.exists() and path.suffix.lower() == ".xlsx":
                self.excel_path = path
                self.file_label.setText(f"Selected file: {path.name}")
                self.output_label.setText(
                    f"Output folder: {path.parent / DEFAULT_OUTPUT_DIR}/"
                )
                self.output_dir = path.parent / DEFAULT_OUTPUT_DIR
                self.clear_button.setEnabled(True)

                try:
                    rows, code_index, image_index = load_rows(path)
                    self.total_images = count_image_urls(path)
                    self.log(f"Loaded: {path.name} ({self.total_images} images)", "info")
                    self.start_button.setEnabled(True)
                except Exception as e:
                    self.log(f"Error: {str(e)}", "error")

        def on_browse(self):
            from PyQt6.QtWidgets import QFileDialog

            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select Excel File", "", "Excel Files (*.xlsx)"
            )
            if file_path:
                self.on_file_dropped(file_path)

        def on_clear(self):
            self.file_label.setText("Selected file: (none)")
            self.output_label.setText("Output folder: downloaded_images/")
            self.clear_button.setEnabled(False)
            self.start_button.setEnabled(False)
            self.excel_path = None
            self.output_dir = None
            self.total_images = 0

        def on_start(self):
            if self.download_thread and self.download_thread.isRunning():
                self.download_thread.stop()
                self.start_button.setText("▶  START DOWNLOAD")
                self.start_button.setEnabled(True)
                return

            if not hasattr(self, "excel_path") or not self.excel_path:
                return

            self.start_button.setText("■  STOP")
            self.start_button.setEnabled(True)
            self.start_button.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                    padding: 12px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #da190b;
                }
            """)
            self.browse_button.setEnabled(False)
            self.clear_button.setEnabled(False)

            self.start_time = time.time()
            self.download_thread = DownloadThread(self.excel_path, self.output_dir)
            self.download_thread.log_signal.connect(self.on_log)
            self.download_thread.progress_signal.connect(self.on_progress)
            self.download_thread.finished_signal.connect(self.on_finished)
            self.download_thread.start()

            self.log("Starting download...", "info")

        def on_log(self, message: str, level: str):
            if level == "success":
                color = "#4CAF50"
            elif level == "error":
                color = "#f44336"
            elif level == "warning":
                color = "#FF9800"
            else:
                color = "#d4d4d4"

            self.log_viewer.append(f'<span style="color: {color};">{message}</span>')
            self.log_viewer.verticalScrollBar().setValue(
                self.log_viewer.verticalScrollBar().maximum()
            )

        def on_progress(self, current: int, total: int):
            elapsed = int(time.time() - self.start_time) if self.start_time else 0
            mins, secs = divmod(elapsed, 60)
            time_str = f"{mins:02d}:{secs:02d}"
            self.status_label.setText(f"Status: {current}/{total} images | {time_str} elapsed")

        def on_finished(self, success: int, failed: int, total: int):
            elapsed = int(time.time() - self.start_time) if self.start_time else 0
            mins, secs = divmod(elapsed, 60)
            time_str = f"{mins:02d}:{secs:02d}"
            self.status_label.setText(
                f"Status: {success}/{total} downloaded ({failed} failed) | {time_str} total"
            )

            self.log(f"Completed: {success} success, {failed} failed", "info")

            self.start_button.setText("▶  START DOWNLOAD")
            self.start_button.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                    padding: 12px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            self.browse_button.setEnabled(True)
            self.clear_button.setEnabled(True)

        def log(self, message: str, level: str):
            self.on_log(message, level)

    window = MainWindow()
    window.show()
    return app.exec()


if __name__ == "__main__":
    sys.exit(main())